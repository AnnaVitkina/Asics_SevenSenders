"""Process AccessorialCosts, Surcharges, and related fields into Excel output."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ACCESSORIAL_SHEET_TITLE = "Accessorials & Surcharges"
DEFAULT_APPLIES_IF = "Invoiced by Carrier"
DEFAULT_RATE_BY_SHIPMENT = "per shipment"
DEFAULT_RATE_BY_WEIGHT = "Weight/kg"

# RA name -> RC name (search extracted costs against either column)
RA_TO_RC_MAPPING: dict[str, str] = {
    "Bulky Delivery To Floor Fee": "Bulky Delivery to Floor",
    "Delivery To Supermarkets Fee": "Delivery to Supermarkets/Large Delivery Shops",
    "Difficulty Surcharge": "Difficult Access",
    "Islands Fee": "Islands",
    "Remote Area Surcharge": "Remote Area",
    "Warehouse Deliveries By Appointment Fee": "Warehouse Deliveries by Appointment",
    "Home Delivery B2B Fee": "Home Delivery With B2B Account",
    "Manual Handling Fee": "Manual Handling",
    "Non-Compliant Parcels 1": "Non-Compliant Parcels 1",
    "Non-Compliant Parcels 2": "Non-Compliant Parcels 2",
    "Out-Of-Area Parcels Fee": "Out-Of-Area Parcels",
    "Oversize Surcharge": "Oversized",
    "Pudo Notification By Sms Fee": "Pudo Notification By Sms",
    "Relabeling Surcharge": "Relabeling",
    "Relabeling Surcharge (PostNord)": "Relabeling",
    "Shipment Data Error Fee": "Shipment Data Error",
    "Forwarding Request": "Forwarding Request",
    "Parcel Volume 100L Fee": "Parcel Volume > 100L",
    "Parcel Volume 200L Fee": "Parcel Volume > 100L",
    "Peak Surcharge": "Peak Surcharge",
    "Peak Surcharge (PostNord)": "Peak Surcharge",
    "Address Clarification Fee": "Address Clarification",
    "Islands 1 Fee": "Islands 1",
    "Islands 10 Fee": "Islands 10",
    "Islands 11 Fee": "Islands 11",
    "Islands 2 Fee": "Islands 2",
    "Islands 4 Fee": "Islands 4",
    "Islands 5 Fee": "Islands 5",
    "Missing Consumer Contact Info Fee": "Missing Consumer Contact Info",
    "Missing Consumer Contact Info Fee (PostNord)": "Missing Consumer Contact Info",
    "Remote Area 1 Fee": "Remote Area 1",
    "Remote Area 2 Fee": "Remote Area 2",
    "Oversize 1 Surcharge": "Oversized 1",
    "Oversize 2 Surcharge": "Oversized 2",
    "Proof of Delivery": "Proof of Delivery Accessibility / Proof of Delivery",
    "Additional Insurance": "Additional Insurance",
    "Consignee Notification Fee": "Consignee notification",
    "Paperless return": "Paperless return",
    "Consignee Only Fee": "Consignee Only",
    "Monday Delivery": "Monday Delivery",
    "Proof of Delivery (Signature)": "Proof of Delivery (signature) + Consignee only",
    "Additional Delivery Attempt": "Additional Delivery Attempt",
    "Returns Processing": "Returns Processing",
    "Special Repatriation": "Special Repatriation",
    "7S Cross-Dock Relabeling": "Relabeling",
    "Return of Non-7S Parcels": "Return of Non-7S Parcels",
    "7S Cross-Dock Shipment Data Error Fee": "Shipment Data Error",
    "Missing customs data": "Missing customs data",
    "Delivery Stop": "Delivery Stop",
    "Pallet Exchange Fee": "4. EURO Pallet exchange BRT IT",
    "EURO Pallet exchange BRT IT": "4. EURO Pallet exchange BRT IT",
    "Canaries": "Canaries (except Hierro & Gomera)",
    "Canaries (except Hierro & Gomera)": "Canaries (except Hierro & Gomera)",
    "IMO": "additional IMO / sea transport surcharge",
    "Secured Parking Fee": "Security Parking Fee (SEUR)",
    "Security Parking Fee (SEUR)": "Security Parking Fee (SEUR)",
    "Proof of Delivery Accessibility": "Proof of Delivery Accessibility / Proof of Delivery",
}

RATE_CARD_ACCESSORIAL_NAMES = {
    "insurance price per 7s shipment",
    "liability (see also 4.2)",
    "7sgreen - co2 offset delivery",
}

RATE_CARD_SURCHARGE_KEYS = {
    ("brt it", "storage"),
    ("dpd fr", "absent consignee"),
    ("post nl", "parcel volume > 50l"),
}

ACCESSORIAL_TAB_HEADERS = [
    "Cost name from RC",
    "Cost name from RA",
    "Price",
    "Currency",
    "Rate by",
    "Apply if",
]


@dataclass
class RateCardExtraColumn:
    title: str
    applies_if: str
    rate_by: str
    currency: str
    value: float | str
    partner_filter: str | None = None


@dataclass
class AccessorialRow:
    rc_name: str
    ra_name: str
    price: float | str
    currency: str
    rate_by: str
    applies_if: str
    source: str


@dataclass
class AccessorialReport:
    rows: list[AccessorialRow] = field(default_factory=list)
    unmapped_extracted: list[str] = field(default_factory=list)
    mapping_not_found: list[str] = field(default_factory=list)
    price_mismatches: list[str] = field(default_factory=list)


def merge_accessorial_reports(reports: list[AccessorialReport]) -> AccessorialReport:
    merged = AccessorialReport()
    seen_rows: set[str] = set()
    for report in reports:
        for row in report.rows:
            key = (
                f"{row.rc_name}|{row.ra_name}|{row.price}|{row.applies_if}|{row.source}"
            )
            if key in seen_rows:
                continue
            seen_rows.add(key)
            merged.rows.append(row)
        merged.unmapped_extracted.extend(report.unmapped_extracted)
        merged.price_mismatches.extend(report.price_mismatches)
    mapping_seen: set[str] = set()
    for report in reports:
        for entry in report.mapping_not_found:
            if entry not in mapping_seen:
                mapping_seen.add(entry)
                merged.mapping_not_found.append(entry)
    return merged


def normalize_match_key(name: str) -> str:
    text = re.sub(r"\s+", " ", (name or "").replace("\n", " ").strip().lower())
    return text


def build_mapping_lookups() -> tuple[dict[str, tuple[str, str]], dict[str, tuple[str, str]]]:
    ra_lookup: dict[str, tuple[str, str]] = {}
    rc_lookup: dict[str, tuple[str, str]] = {}
    for ra, rc in RA_TO_RC_MAPPING.items():
        if not rc:
            continue
        ra_lookup[normalize_match_key(ra)] = (ra, rc)
        rc_lookup[normalize_match_key(rc)] = (ra, rc)
        for part in re.split(r"\s*/\s*", rc):
            part = part.strip()
            if part:
                rc_lookup[normalize_match_key(part)] = (ra, rc)
    return ra_lookup, rc_lookup


def resolve_mapping(name: str) -> tuple[str, str] | None:
    ra_lookup, rc_lookup = build_mapping_lookups()
    key = normalize_match_key(clean_display_name(name))
    if key in ra_lookup:
        return ra_lookup[key]
    if key in rc_lookup:
        return rc_lookup[key]
    for lookup in (ra_lookup, rc_lookup):
        for norm_key, pair in lookup.items():
            if norm_key in key or key in norm_key:
                return pair
    return None


def clean_display_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").split("\n")[0].strip())


def parse_currency_from_text(text: str) -> str:
    for sym, code in (("€", "EUR"), ("$", "USD"), ("£", "GBP")):
        if sym in text or code in text.upper():
            return code
    return "EUR"


def parse_price_value(text: str) -> float | str | None:
    if not text:
        return None
    raw = str(text).strip()
    if "%" in raw and "of Rate" in raw:
        return raw
    m = re.search(
        r"(?:EUR|€)\s*([\d.,]+)|([\d.,]+)\s*(?:EUR|€)|(?:^|[\s+])([\d.,]+)\s*(?:EUR|€|/kg|per)",
        raw,
        re.I,
    )
    if m:
        num = (m.group(1) or m.group(2) or m.group(3) or "").replace(",", ".")
        try:
            return float(num)
        except ValueError:
            pass
    m = re.search(r"([\d]+[.,][\d]+|\d+)", raw)
    if m:
        try:
            return float(m.group(1).replace(",", "."))
        except ValueError:
            pass
    return None


def infer_rate_by(price_text: str, info_text: str = "") -> str:
    combined = f"{price_text} {info_text}".lower()
    if "/kg" in combined or "per kg" in combined:
        return DEFAULT_RATE_BY_WEIGHT
    return DEFAULT_RATE_BY_SHIPMENT


def build_applies_if(
    lmc: str | None = None, country_applicable: str | None = None
) -> str:
    parts = [DEFAULT_APPLIES_IF]
    if lmc and str(lmc).strip() not in ("", "-"):
        parts.append(f"Carrier Partner equals {lmc.strip()}")
    country = (country_applicable or "").strip()
    if country and country != "-" and re.fullmatch(r"[A-Z]{2}", country):
        parts.append(f"Destination Country equals {country}")
    return "; ".join(parts)


def normalize_partner(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().upper())


def partner_matches(lane_partner: str, filter_partner: str | None) -> bool:
    if not filter_partner:
        return True
    lane_p = normalize_partner(lane_partner)
    filter_p = normalize_partner(filter_partner)
    if lane_p == filter_p:
        return True
    if filter_p.startswith(lane_p + " ") or filter_p.startswith(lane_p):
        return True
    if lane_p == filter_p.split()[0]:
        return True
    return False


def _first_accessorial_by_name(
    items: list[dict], cost_name: str
) -> dict | None:
    target = normalize_match_key(cost_name)
    for item in items:
        if normalize_match_key(item.get("CostName", "")) == target:
            return item
    return None


def build_rate_card_extra_columns(fields: dict) -> tuple[list[RateCardExtraColumn], list[str]]:
    """Build fixed extra columns for the Rate Card sheet."""
    accessorial = fields.get("AccessorialCosts") or []
    surcharges = fields.get("Surcharges") or []
    mismatches: list[str] = []

    insurance_item = _first_accessorial_by_name(
        accessorial, "Insurance Price per 7S Shipment"
    )
    liability_item = _first_accessorial_by_name(accessorial, "Liability (see also 4.2)")

    insurance_price = None
    liability_price = None
    if insurance_item:
        insurance_price = parse_price_value(insurance_item.get("Information", ""))
    if liability_item:
        liability_price = parse_price_value(liability_item.get("Information", ""))

    if (
        insurance_price is not None
        and liability_price is not None
        and isinstance(insurance_price, (int, float))
        and isinstance(liability_price, (int, float))
        and abs(insurance_price - liability_price) > 0.001
    ):
        mismatches.append(
            "Claims Center And Insurance Fee: Insurance Price per 7S Shipment "
            f"({insurance_price}) differs from Liability (see also 4.2) ({liability_price})"
        )

    claim_value = insurance_price if insurance_price is not None else liability_price
    if claim_value is None:
        claim_value = 0.17

    columns: list[RateCardExtraColumn] = [
        RateCardExtraColumn(
            title="Claims Center And Insurance Fee",
            applies_if=DEFAULT_APPLIES_IF,
            rate_by=DEFAULT_RATE_BY_SHIPMENT,
            currency="EUR",
            value=claim_value,
        ),
    ]

    green = _first_accessorial_by_name(accessorial, "7SGreen - CO2 offset Delivery")
    green_price = (
        parse_price_value(green.get("Information", "")) if green else 0.03
    )
    columns.append(
        RateCardExtraColumn(
            title="7S Green Surcharge",
            applies_if=DEFAULT_APPLIES_IF,
            rate_by=DEFAULT_RATE_BY_SHIPMENT,
            currency="EUR",
            value=green_price if green_price is not None else 0.03,
        )
    )

    storage = _find_surcharge(surcharges, "BRT IT", "Storage")
    storage_price = parse_price_value(storage.get("Price", "")) if storage else 4.0
    columns.append(
        RateCardExtraColumn(
            title="Storage Fee",
            applies_if=f"{DEFAULT_APPLIES_IF}; Carrier Partner equals BRT IT",
            rate_by=DEFAULT_RATE_BY_SHIPMENT,
            currency=parse_currency_from_text(storage.get("Price", "") if storage else "€"),
            value=storage_price if storage_price is not None else 4.0,
            partner_filter="BRT IT",
        )
    )

    absent = _find_surcharge(surcharges, "DPD FR", "Absent consignee")
    absent_price = parse_price_value(absent.get("Price", "")) if absent else 3.0
    columns.append(
        RateCardExtraColumn(
            title="Absent consignee",
            applies_if=f"{DEFAULT_APPLIES_IF}; Carrier Partner equals DPD FR",
            rate_by=DEFAULT_RATE_BY_SHIPMENT,
            currency=parse_currency_from_text(absent.get("Price", "") if absent else "€"),
            value=absent_price if absent_price is not None else 3.0,
        )
    )

    parcel = _find_surcharge(surcharges, "Post NL", "Parcel Volume > 50L")
    parcel_price = parse_price_value(parcel.get("Price", "")) if parcel else 1.0
    columns.append(
        RateCardExtraColumn(
            title="Parcel Volume 50L Fee",
            applies_if=f"{DEFAULT_APPLIES_IF}; Carrier Partner equals Post NL",
            rate_by=DEFAULT_RATE_BY_SHIPMENT,
            currency=parse_currency_from_text(parcel.get("Price", "") if parcel else "€"),
            value=parcel_price if parcel_price is not None else 1.0,
        )
    )

    return columns, mismatches


def _find_surcharge(
    surcharges: list[dict], lmc: str, name: str
) -> dict | None:
    lmc_n = normalize_partner(lmc)
    name_n = normalize_match_key(name)
    for row in surcharges:
        if normalize_partner(row.get("LMC", "")) != lmc_n:
            continue
        if normalize_match_key(row.get("Name", "")) == name_n:
            return row
    return None


def collect_extracted_items(fields: dict) -> list[dict]:
    items: list[dict] = []
    for row in fields.get("AccessorialCosts") or []:
        cost_name = (row.get("CostName") or "").strip()
        if not cost_name:
            continue
        items.append(
            {
                "name": cost_name,
                "price_text": row.get("Information", ""),
                "source": "AccessorialCosts",
                "lmc": None,
                "country": None,
            }
        )

    for row in fields.get("Surcharges") or []:
        name = clean_display_name(row.get("Name", ""))
        if not name:
            continue
        items.append(
            {
                "name": name,
                "price_text": row.get("Price", ""),
                "source": "Surcharges",
                "lmc": row.get("LMC"),
                "country": row.get("CountryApplicable"),
            }
        )

    for row in fields.get("Additionals") or []:
        cost = (row.get("Cost") or "").strip()
        if not cost:
            continue
        items.append(
            {
                "name": cost,
                "price_text": row.get("Details", ""),
                "source": "Additionals",
                "lmc": None,
                "country": None,
            }
        )

    for row in fields.get("SeaRates") or []:
        region = (row.get("Region") or "").strip()
        if not region or region.lower() == "region":
            continue
        items.append(
            {
                "name": region,
                "price_text": row.get("Surcharge", ""),
                "source": "SeaRates",
                "lmc": "SEUR ES",
                "country": "ES",
            }
        )

    return items


def _is_rate_card_only_item(item: dict) -> bool:
    if item["source"] == "AccessorialCosts":
        return normalize_match_key(item["name"]) in RATE_CARD_ACCESSORIAL_NAMES
    if item["source"] == "Surcharges":
        key = (
            normalize_partner(item.get("lmc") or ""),
            normalize_match_key(item["name"]),
        )
        return key in RATE_CARD_SURCHARGE_KEYS
    return False


def process_accessorial_data(fields: dict) -> AccessorialReport:
    report = AccessorialReport()
    matched_rc_keys: set[str] = set()
    seen_tab_keys: set[str] = set()

    for item in collect_extracted_items(fields):
        name = item["name"]
        if _is_rate_card_only_item(item):
            continue

        mapping = resolve_mapping(name)
        if mapping is None:
            report.unmapped_extracted.append(
                f"[{item['source']}] {name} | {item['price_text'][:80]}"
            )
            continue

        ra_name, rc_name = mapping
        matched_rc_keys.add(normalize_match_key(rc_name))

        price = parse_price_value(item["price_text"])
        if price is None:
            price = item["price_text"] or ""

        tab_key = (
            f"{rc_name}|{ra_name}|{item.get('lmc')}|{item.get('country')}|{price}"
        )
        if tab_key in seen_tab_keys:
            continue
        seen_tab_keys.add(tab_key)

        report.rows.append(
            AccessorialRow(
                rc_name=rc_name,
                ra_name=ra_name,
                price=price,
                currency=parse_currency_from_text(item["price_text"]),
                rate_by=infer_rate_by(item["price_text"]),
                applies_if=build_applies_if(item.get("lmc"), item.get("country")),
                source=item["source"],
            )
        )

    for ra_name, rc_name in RA_TO_RC_MAPPING.items():
        if not rc_name:
            continue
        if normalize_match_key(rc_name) not in matched_rc_keys:
            report.mapping_not_found.append(f"{ra_name} -> {rc_name}")

    _, extra_mismatches = build_rate_card_extra_columns(fields)
    report.price_mismatches.extend(extra_mismatches)

    return report


def write_accessorial_sheet(workbook: Workbook, report: AccessorialReport) -> None:
    ws = workbook.create_sheet(ACCESSORIAL_SHEET_TITLE[:31])
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(ACCESSORIAL_TAB_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, entry in enumerate(report.rows, 2):
        ws.cell(row=row_idx, column=1, value=entry.rc_name)
        ws.cell(row=row_idx, column=2, value=entry.ra_name)
        ws.cell(row=row_idx, column=3, value=entry.price)
        ws.cell(row=row_idx, column=4, value=entry.currency)
        ws.cell(row=row_idx, column=5, value=entry.rate_by)
        ws.cell(row=row_idx, column=6, value=entry.applies_if)

    last_row = max(1, len(report.rows) + 1)
    for c in range(1, len(ACCESSORIAL_TAB_HEADERS) + 1):
        letter = get_column_letter(c)
        max_len = len(ACCESSORIAL_TAB_HEADERS[c - 1])
        for r in range(1, last_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    if report.rows:
        ws.auto_filter.ref = f"A1:F{last_row}"


def write_report_files(report: AccessorialReport, out_path: Path) -> None:
    stem = out_path.stem
    parent = out_path.parent

    if report.unmapped_extracted:
        path = parent / f"{stem}_accessorial_unmapped.txt"
        path.write_text(
            "Extracted costs not in RA/RC mapping:\n\n"
            + "\n".join(report.unmapped_extracted),
            encoding="utf-8",
        )

    if report.mapping_not_found:
        path = parent / f"{stem}_accessorial_mapping_not_found.txt"
        path.write_text(
            "Mapping entries not found in extracted data:\n\n"
            + "\n".join(report.mapping_not_found),
            encoding="utf-8",
        )

    if report.price_mismatches:
        path = parent / f"{stem}_accessorial_price_mismatch.txt"
        path.write_text(
            "Price mismatches:\n\n" + "\n".join(report.price_mismatches),
            encoding="utf-8",
        )


def _style_header_cell(cell, header_fill, header_font, header_alignment) -> None:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment


def write_simple_cost_block_headers(
    ws,
    start_col: int,
    column: RateCardExtraColumn,
    header_fill,
    header_font,
    header_alignment,
) -> int:
    end_col = start_col + 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=column.title)
    _style_header_cell(cell, header_fill, header_font, header_alignment)

    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    cell = ws.cell(row=2, column=start_col, value=column.applies_if)
    _style_header_cell(cell, header_fill, header_font, header_alignment)

    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
    cell = ws.cell(row=3, column=start_col, value=f"Rate by: {column.rate_by}")
    _style_header_cell(cell, header_fill, header_font, header_alignment)

    ws.cell(row=4, column=start_col, value="")
    ws.cell(row=4, column=start_col).fill = header_fill
    ws.cell(row=4, column=start_col + 1, value="")
    ws.cell(row=4, column=start_col + 1).fill = header_fill

    cell = ws.cell(row=5, column=start_col, value="Currency")
    _style_header_cell(cell, header_fill, header_font, header_alignment)
    cell = ws.cell(row=5, column=start_col + 1, value="Flat")
    _style_header_cell(cell, header_fill, header_font, header_alignment)
    return end_col
