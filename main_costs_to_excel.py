"""Build a Rate Card Excel sheet from cleaned MainCosts JSON."""

from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROCESSING_DIR = Path(__file__).resolve().parent / "processing"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"
SHEET_TITLE = "Rate Card"

FIXED_COLS = [
    "CARRIER",
    "Carrier Partner",
    "Origin country",
    "Origin Postal Code",
    "Destination Country",
    "Service type",
    "Valid From",
    "Valid to",
]

DEFAULT_RATE_BY_WEIGHT = "Weight/kg"
FLOATER_RATE_BY = "Condition/ParcelPallet"

ORIGIN_SITES = {
    "krefeld": ("DE", "47807", "Krefeld"),
    "garons": ("FR", "30128", "Garons"),
}

COUNTRY_TO_CODE = {
    "spain": "ES",
    "portugal": "PT",
    "france": "FR",
    "italy": "IT",
    "netherlands": "NL",
    "belgium": "BE",
    "sweden": "SE",
    "denmark": "DK",
    "finland": "FI",
    "norway": "NO",
    "germany": "DE",
    "austria": "AT",
    "switzerland": "CH",
    "poland": "PL",
    "czech": "CZ",
    "czechia": "CZ",
    "hungary": "HU",
    "romania": "RO",
    "bulgaria": "BG",
    "croatia": "HR",
    "slovenia": "SI",
    "slovakia": "SK",
    "estonia": "EE",
    "latvia": "LV",
    "lithuania": "LT",
    "luxembourg": "LU",
    "ireland": "IE",
    "greece": "GR",
}

CURRENCY_SYMBOLS = {"€": "EUR", "$": "USD", "£": "GBP"}

SKIP_SECTION_MARKERS = (
    "packing density",
    "floater",
    "discount / surcharge",
    "shipments per pallet",
    "parcels per pallet",
)

SERVICE_PATTERNS = [
    (re.compile(r"\bRETURN\b", re.I), "Return"),
    (re.compile(r"\bHOME\b", re.I), "HOME"),
    (re.compile(r"\bPUDO\b", re.I), "PUDO"),
    (re.compile(r"\bB2B\b", re.I), "B2B"),
    (re.compile(r"\bB2C\b", re.I), "B2C"),
    (re.compile(r"\bCLASSIC\b", re.I), "Classic"),
    (re.compile(r"\bUNDELIVERABLES\b", re.I), "Undeliverables"),
]

PARTNER_ALIASES = {
    "POSTE ITALIANE": "SDA IT",
}


@dataclass
class OriginInfo:
    country: str
    postal_code: str
    city: str
    label: str  # e.g. "Ex Krefeld"


@dataclass
class PriceBlock:
    partner: str
    section_title: str
    section_country: str | None
    origin: OriginInfo
    destination_code: str | None
    headers: dict[str, str]
    weight_rows: list[dict[str, str]]


@dataclass
class LaneRow:
    carrier: str
    carrier_partner: str
    origin_country: str
    origin_postal: str
    destination_country: str
    service_type: str
    valid_from: str
    valid_to: str
    transport_title: str
    currency: str
    brackets: dict[str, object]
    bracket_rate_types: dict[str, str]
    forward_filled: set[str] = field(default_factory=set)
    applies_if: str = ""
    rate_by_label: str = DEFAULT_RATE_BY_WEIGHT


@dataclass
class FloaterColumn:
    title: str
    applies_if: str
    rate_by: str
    currency: str
    brackets: dict[str, object]
    bracket_rate_types: dict[str, str]
    forward_filled: set[str] = field(default_factory=set)


def is_floater_lane(lane: LaneRow) -> bool:
    return lane.transport_title.startswith("Packing Density Floater Fee")


def lane_to_floater_column(lane: LaneRow) -> FloaterColumn:
    return FloaterColumn(
        title=lane.transport_title,
        applies_if=lane.applies_if,
        rate_by=lane.rate_by_label or FLOATER_RATE_BY,
        currency=lane.currency,
        brackets=dict(lane.brackets),
        bracket_rate_types=dict(lane.bracket_rate_types),
        forward_filled=set(lane.forward_filled),
    )


def split_transport_and_floater(
    lanes: list[LaneRow],
) -> tuple[list[LaneRow], list[FloaterColumn]]:
    transport: list[LaneRow] = []
    floaters: list[FloaterColumn] = []
    for lane in lanes:
        if is_floater_lane(lane):
            floaters.append(lane_to_floater_column(lane))
        else:
            transport.append(lane)
    return transport, floaters


def normalize_cost_key(key: str) -> str:
    """Map Cost5, cost5, Ccost5 (OCR typo) to canonical Cost5."""
    m = re.match(r"^c*cost(\d+)$", key, re.I)
    if m:
        return f"Cost{m.group(1)}"
    return key


def is_cost_field(key: str) -> bool:
    return normalize_cost_key(key).startswith("Cost")


def cost_keys(row: dict) -> list[str]:
    keys = [normalize_cost_key(k) for k in row if is_cost_field(k)]
    return sorted(set(keys), key=lambda k: int(re.search(r"\d+", k).group()))


def row_get_cost(row: dict, cost_key: str) -> str | None:
    """Read a cost cell; matches Ccost5 when cost_key is Cost5."""
    if cost_key in row:
        return row.get(cost_key)
    for key, value in row.items():
        if normalize_cost_key(key) == cost_key:
            return value
    return None


def is_basic_fees(name: str) -> bool:
    return name.strip().lower().startswith("basic fees")


def is_packing_density_floater_start(name: str) -> bool:
    low = name.lower()
    return "packing density" in low and "floater" in low


def is_floater_region_name(name: str) -> bool:
    low = name.strip().lower()
    return low in (
        "spain & portugal",
        "italy",
        "france",
    )


def should_skip_section(name: str) -> bool:
    if is_packing_density_floater_start(name):
        return False
    low = name.lower()
    return any(marker in low for marker in SKIP_SECTION_MARKERS)


def normalize_carrier_label(carrier_field: str) -> str:
    text = (carrier_field or "Seven Senders").strip()
    if " or " in text.lower():
        return text.split(" or ")[0].strip()
    return text or "Seven Senders"


def is_floater_header_row(row: dict) -> bool:
    if not cost_keys(row):
        return False
    weight = (row.get("Weight") or "").lower()
    return (
        "discount" in weight
        or "pallet" in weight
        or "pick-up" in weight
        or "pick up" in weight
    )


def is_floater_weight_row(row: dict) -> bool:
    weight = (row.get("Weight") or "").strip()
    return weight.startswith("≥") or weight.startswith(">=")


def parse_floater_threshold(weight: str) -> float | None:
    w = weight.strip()
    m = re.match(r"(?:≥|>=|>)\s*([\d.,]+)", w)
    if not m:
        return None
    return float(m.group(1).replace(",", "."))


def floater_column_label_less(threshold: float) -> str:
    if threshold == int(threshold):
        return f"< {int(threshold)}"
    return f"< {threshold}"


def floater_column_label_greater_equal(threshold: float) -> str:
    if threshold == int(threshold):
        return f"≥ {int(threshold)}"
    return f"≥ {threshold}"


def floater_brackets_from_rows(
    weight_rows: list[dict[str, str]], cost_key: str
) -> tuple[dict[str, object], dict[str, str], set[str]]:
    """Map ≥ threshold rows to <N columns (prior tier value) and a final ≥N column."""
    tiers: list[tuple[float, float]] = []
    for row in weight_rows:
        threshold = parse_floater_threshold(row.get("Weight") or "")
        cost = parse_cost_value(row_get_cost(row, cost_key))
        if threshold is None or cost is None:
            continue
        if tiers and tiers[-1][0] == threshold:
            tiers[-1] = (threshold, cost)
        else:
            tiers.append((threshold, cost))

    if not tiers:
        return {}, {}, set()

    tiers.sort(key=lambda item: item[0])
    brackets: dict[str, object] = {}
    rate_types: dict[str, str] = {}

    for i, (threshold, _) in enumerate(tiers):
        label = floater_column_label_less(threshold)
        value = tiers[i - 1][1] if i > 0 else tiers[0][1]
        brackets[label] = value
        rate_types[label] = "Flat"

    last_threshold, last_cost = tiers[-1]
    ge_label = floater_column_label_greater_equal(last_threshold)
    brackets[ge_label] = last_cost
    rate_types[ge_label] = "Flat"

    return brackets, rate_types, set()


def floater_applies_if(region_name: str) -> str:
    mapping = {
        "Spain & Portugal": "Destination Country equals ES, PT",
        "Italy": "Destination Country equals IT",
        "France": "Destination Country equals FR",
    }
    return mapping.get(region_name.strip(), f"Destination Country equals {region_name}")


def floater_destination_country(region_name: str) -> str:
    mapping = {
        "Spain & Portugal": "ES, PT",
        "Italy": "IT",
        "France": "FR",
    }
    return mapping.get(region_name.strip(), "")


def floater_transport_title(region_name: str) -> str:
    return f"Packing Density Floater Fee ({region_name})"


def build_floater_lane(
    region_name: str,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    carrier_label: str,
    valid_from: str,
    valid_to: str,
) -> LaneRow | None:
    cost_key = "Cost1"
    brackets, rate_types, forward_filled = floater_brackets_from_rows(
        weight_rows, cost_key
    )
    if not brackets:
        return None
    sample = next(iter(brackets.values()))
    dest = floater_destination_country(region_name)
    return LaneRow(
        carrier=f"{carrier_label} {dest}".strip(),
        carrier_partner=carrier_label,
        origin_country=origin.country,
        origin_postal=origin.postal_code,
        destination_country=dest,
        service_type="Packing Density Floater",
        valid_from=valid_from,
        valid_to=valid_to,
        transport_title=floater_transport_title(region_name),
        currency=parse_currency("€/Parcel", str(sample)),
        brackets=brackets,
        bracket_rate_types=rate_types,
        forward_filled=forward_filled,
        applies_if=floater_applies_if(region_name),
        rate_by_label=FLOATER_RATE_BY,
    )


def process_floater_merged_sp_it(
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    carrier_label: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    lanes: list[LaneRow] = []
    sp_rows = weight_rows[0::2]
    it_rows = weight_rows[1::2]
    for region_name, rows in (
        ("Spain & Portugal", sp_rows),
        ("Italy", it_rows),
    ):
        lane = build_floater_lane(
            region_name, rows, origin, carrier_label, valid_from, valid_to
        )
        if lane:
            lanes.append(lane)
    return lanes


def process_floater_single_region(
    region_name: str,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    carrier_label: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    lane = build_floater_lane(
        region_name, weight_rows, origin, carrier_label, valid_from, valid_to
    )
    return [lane] if lane else []


def is_origin_rate_name(name: str) -> bool:
    low = name.lower().replace("\n", " ")
    return (
        "krefeld" in low
        or "garons" in low
        or low.startswith("base prices")
        or low.startswith("ex ")
        or low.startswith("to ")
    )


def is_country_rate_name(name: str) -> bool:
    return name.strip().lower().startswith("country")


def _is_service_header_text(value: str) -> bool:
    text = value.strip()
    if not text or parse_cost_value(text) is not None:
        return False
    low = text.lower()
    return any(kw in low for kw in ("shipping", "return", "undeliverable", "parcel", "shipment"))


def is_price_header_row(row: dict) -> bool:
    weight = (row.get("Weight") or "").strip()
    keys = cost_keys(row)
    if not keys:
        return False
    if weight == "Weight (up to)":
        return True
    if is_noise_weight(weight):
        return False
    _, numeric_weight = parse_weight_label(weight)
    if numeric_weight is not None:
        return False
    return any(_is_service_header_text(str(row.get(k) or "")) for k in keys)


def is_noise_weight(weight: str) -> bool:
    w = weight.strip()
    if not w or w == "Weight (up to)":
        return True
    if w.startswith("≥") or w.startswith(">="):
        return True
    if "pallet" in w.lower():
        return True
    if "discount" in w.lower():
        return True
    return False


def parse_origin(name: str, row: dict | None = None) -> OriginInfo | None:
    text = name.replace("\n", " ")
    if row and row.get("RateName"):
        text = f"{text} {row['RateName']}".replace("\n", " ")
    low = text.lower()
    if "krefeld" in low:
        country, postal, city = ORIGIN_SITES["krefeld"]
        return OriginInfo(country, postal, city, "Ex Krefeld")
    if "garons" in low:
        country, postal, city = ORIGIN_SITES["garons"]
        label = "To Garons" if "to garons" in low and "ex garons" not in low else "Ex Garons"
        return OriginInfo(country, postal, city, label)
    return None


def parse_partner_from_section(title: str) -> str:
    m = re.search(r"\(([^)]+)\)\s*$", title)
    if m:
        return m.group(1).strip()
    return ""


def parse_section_country(title: str) -> str | None:
    low = title.lower()
    if "nordic" in low:
        return None
    if "spain" in low and "portugal" in low:
        return None
    if "netherlands" in low and "belgium" in low:
        return None
    for country, code in COUNTRY_TO_CODE.items():
        if country in low:
            return code
    return None


def parse_country_codes_from_rate_name(name: str) -> list[str]:
    body = re.sub(r"(?i)^country\s*", "", name.strip()).strip()
    if not body:
        return []
    parts = [p.strip().rstrip("*") for p in body.split("\n") if p.strip()]
    codes: list[str] = []
    for part in parts:
        low = part.lower()
        found = False
        for country, code in COUNTRY_TO_CODE.items():
            if country in low:
                codes.append(code)
                found = True
                break
        if not found and len(part) == 2 and part.isalpha():
            codes.append(part.upper())
    return codes


def is_postnord_partner(partner: str, section_title: str = "") -> bool:
    return "postnord" in (partner + section_title).lower()


def has_multiline_header(row: dict) -> bool:
    for key in cost_keys(row):
        val = str(row.get(key) or "")
        if "\n" in val and _is_service_header_text(val.split("\n")[0]):
            return True
    return False


def has_paired_weight_rows(weight_rows: list[dict]) -> bool:
    from itertools import groupby

    for _weight, group in groupby(weight_rows, key=lambda r: r.get("Weight", "")):
        rows = list(group)
        if len(rows) >= 2:
            return True
    return False


def split_header_by_country(row: dict) -> dict[str, list[str]]:
    split: dict[str, list[str]] = {}
    for key in cost_keys(row):
        val = str(row.get(key) or "")
        lines = [ln.strip() for ln in val.split("\n") if ln.strip()]
        split[key] = lines if lines else [val]
    return split


def group_weight_row_pairs(
    weight_rows: list[dict[str, str]],
) -> list[tuple[str, list[dict[str, str]]]]:
    groups: list[tuple[str, list[dict[str, str]]]] = []
    i = 0
    while i < len(weight_rows):
        weight = weight_rows[i].get("Weight", "")
        batch = [weight_rows[i]]
        i += 1
        while i < len(weight_rows) and weight_rows[i].get("Weight") == weight:
            batch.append(weight_rows[i])
            i += 1
        groups.append((weight, batch))
    return groups


def weight_rows_for_country_index(
    weight_rows: list[dict[str, str]], country_index: int
) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for weight, batch in group_weight_row_pairs(weight_rows):
        if country_index < len(batch):
            row = dict(batch[country_index])
            row["Weight"] = weight
            result.append(row)
    return result


def weight_rows_norway_only(weight_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    result = []
    for weight, batch in group_weight_row_pairs(weight_rows):
        for row in batch:
            c1 = parse_cost_value(row_get_cost(row, "Cost1"))
            c2 = parse_cost_value(row_get_cost(row, "Cost2"))
            if c1 is not None and c2 is None and "Cost2" not in row:
                result.append({**row, "Weight": weight})
            elif len(batch) == 1 and c1 is not None and c2 is None:
                raw_c2 = row_get_cost(row, "Cost2")
                if raw_c2 is None or str(raw_c2).strip() in ("", "-"):
                    result.append({**row, "Weight": weight})
    return result


def weight_rows_se_dk_paired(weight_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for weight, batch in group_weight_row_pairs(weight_rows):
        if len(batch) >= 2:
            for row in batch:
                result.append(dict(row))
        elif len(batch) == 1:
            c2 = parse_cost_value(row_get_cost(batch[0], "Cost2"))
            if c2 is not None:
                result.append(dict(batch[0]))
    return result


def parse_postnord_zone_service(header: str) -> str:
    text = _header_line(header)
    zone_m = re.search(r"Zone\s+(\d+)", text, re.I)
    zone = zone_m.group(1) if zone_m else "?"
    low = text.lower()
    if "return" in low and "pudo" in low:
        return f"Return Zone {zone}"
    if "pudo" in low:
        return f"PUDO Zone {zone}"
    if "home" in low:
        return f"HOME Zone {zone}"
    if "b2b" in low:
        return f"B2B Zone {zone}"
    return f"Zone {zone}"


def postnord_partner_label(country_code: str) -> str:
    return f"PostNord {country_code}"


def build_postnord_carrier(country_code: str) -> str:
    return postnord_partner_label(country_code)


def build_postnord_transport_title(
    country_code: str, max_kg: int | None, has_adder: bool
) -> str:
    kg = max_kg if max_kg is not None else 30
    label = postnord_partner_label(country_code)
    if has_adder:
        return f"Transport Cost ({label} {kg}kg + add kg)"
    return f"Transport Cost ({label} {kg} kg)"


def _make_postnord_lane(
    country_code: str,
    service_type: str,
    header: str,
    cost_key: str,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
) -> LaneRow | None:
    bracket_labels, rate_types, max_kg, has_adder = lane_weight_profile(
        weight_rows, cost_key
    )
    if not bracket_labels:
        return None
    brackets, forward_filled = lane_bracket_values(
        weight_rows, cost_key, bracket_labels
    )
    if not brackets:
        return None
    sample_cost = next(iter(brackets.values()))
    return LaneRow(
        carrier=build_postnord_carrier(country_code),
        carrier_partner=postnord_partner_label(country_code),
        origin_country=origin.country,
        origin_postal=origin.postal_code,
        destination_country=country_code,
        service_type=service_type,
        valid_from=valid_from,
        valid_to=valid_to,
        transport_title=build_postnord_transport_title(country_code, max_kg, has_adder),
        currency=parse_currency(header, str(sample_cost)),
        brackets=brackets,
        bracket_rate_types=rate_types,
        forward_filled=forward_filled,
    )


def _expand_return_undelivered_pair(lane: LaneRow) -> list[LaneRow]:
    out = [lane]
    svc = lane.service_type.upper()
    if svc == "RETURN":
        out.append(clone_lane(lane, "UNDELIVERED"))
    elif svc.startswith("RETURN ZONE"):
        zone = lane.service_type.replace("Return Zone", "").strip()
        out.append(clone_lane(lane, f"Undelivered Zone {zone}"))
    return out


def process_postnord_type1_se_dk(
    header_row: dict,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    """Paired weight rows: first row = SE, second = DK."""
    split = split_header_by_country(header_row)
    lanes: list[LaneRow] = []
    country_codes = ["SE", "DK"]

    for cost_key, header_lines in split.items():
        for idx, header_text in enumerate(header_lines):
            if idx >= len(country_codes):
                break
            country = country_codes[idx]
            wr = weight_rows_for_country_index(weight_rows, idx)
            service = parse_service_type(
                header_text,
                partner="PostNord",
                cost_key=cost_key,
                section_title=section_title,
            )
            lane = _make_postnord_lane(
                country, service, header_text, cost_key, wr, origin, section_title,
                valid_from, valid_to,
            )
            if lane:
                lanes.append(lane)
    return lanes


def process_postnord_type2_single(
    country_code: str,
    header_row: dict,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    headers = extract_headers(header_row)
    lanes: list[LaneRow] = []
    for cost_key, header in headers.items():
        service = parse_service_type(
            header, partner="PostNord", cost_key=cost_key, section_title=section_title
        )
        lane = _make_postnord_lane(
            country_code, service, header, cost_key, weight_rows, origin, section_title,
            valid_from, valid_to,
        )
        if lane:
            lanes.append(lane)
    return lanes


def process_postnord_type3_norway_zones(
    header_row: dict,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    headers = extract_headers(header_row)
    lanes: list[LaneRow] = []
    for cost_key, header in headers.items():
        if "zone" not in header.lower():
            continue
        service = parse_postnord_zone_service(header)
        lane = _make_postnord_lane(
            "NO", service, header, cost_key, weight_rows, origin, section_title,
            valid_from, valid_to,
        )
        if lane:
            lanes.extend(_expand_return_undelivered_pair(lane))
    return lanes


def process_postnord_type4_multi_return(
    header_row: dict,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    headers = extract_headers(header_row)
    lanes: list[LaneRow] = []
    nordic_cols = {"Cost1": "SE", "Cost2": "DK", "Cost3": "FI"}
    shared_return_header = next(
        (h for h in headers.values() if "undeliverable" in h.lower() and "return" in h.lower()),
        "7S Return PUDO / Undeliverables [€/Parcel]",
    )

    for cost_key, header in headers.items():
        if "zone" in header.lower():
            service = parse_postnord_zone_service(header)
            lane = _make_postnord_lane(
                "NO", service, header, cost_key, weight_rows, origin, section_title,
                valid_from, valid_to,
            )
            if lane:
                lanes.extend(_expand_return_undelivered_pair(lane))

    for cost_key, country in nordic_cols.items():
        if not any(parse_cost_value(row_get_cost(r, cost_key)) for r in weight_rows):
            continue
        hdr = headers.get(cost_key, shared_return_header)
        if "zone" in hdr.lower():
            continue
        lane = _make_postnord_lane(
            country,
            "RETURN",
            hdr,
            cost_key,
            weight_rows,
            origin,
            section_title,
            valid_from,
            valid_to,
        )
        if lane:
            lanes.extend(_expand_return_undelivered_pair(lane))
    return lanes


def process_postnord_type5_return_split(
    header_row: dict,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
) -> list[LaneRow]:
    """Cost1 column = SE/DK, Cost2 column = DK; 30 kg-only Cost1 row = Norway."""
    split = split_header_by_country(header_row)
    lanes: list[LaneRow] = []
    cost1_lines = split.get("Cost1", [])
    shared_rows: list[dict[str, str]] = []
    no_rows: list[dict[str, str]] = []

    for weight, batch in group_weight_row_pairs(weight_rows):
        for row in batch:
            r = dict(row)
            r["Weight"] = weight
            c1 = parse_cost_value(row_get_cost(row, "Cost1"))
            c2 = parse_cost_value(row_get_cost(row, "Cost2"))
            raw_c2 = row_get_cost(row, "Cost2")
            if c1 is not None and (c2 is None or raw_c2 is None or str(raw_c2).strip() in ("", "-")):
                if c2 is None and (c1 or 0) > 100:
                    no_rows.append(r)
                    continue
            shared_rows.append(r)

    if len(cost1_lines) >= 2:
        lane_no = _make_postnord_lane(
            "NO",
            "Return B2B",
            cost1_lines[1],
            "Cost1",
            no_rows,
            origin,
            section_title,
            valid_from,
            valid_to,
        )
        if lane_no:
            lanes.append(lane_no)

    header_se = cost1_lines[0] if cost1_lines else "7S Return HOME/B2B [€/Parcel]"
    header_dk = split.get("Cost2", [header_se])[0]

    lane_se = _make_postnord_lane(
        "SE", "Return B2B", header_se, "Cost1", shared_rows, origin, section_title,
        valid_from, valid_to,
    )
    if lane_se:
        lanes.append(lane_se)

    lane_dk = _make_postnord_lane(
        "DK", "Return B2B", header_dk, "Cost2", shared_rows, origin, section_title,
        valid_from, valid_to,
    )
    if lane_dk:
        lanes.append(lane_dk)

    return lanes


def classify_postnord_block(
    country_codes: list[str],
    header_row: dict,
    weight_rows: list[dict[str, str]],
    pending_se: bool,
) -> str:
    headers = extract_headers(header_row)
    headers_text = " ".join(headers.values()).lower()
    codes = set(country_codes)

    if len(codes) >= 3 and "undeliverable" in headers_text and any(
        "zone" in h.lower() for h in headers.values()
    ):
        return "type4"

    if codes == {"NO"} or country_codes == ["NO"]:
        if any("zone" in h.lower() for h in headers.values()):
            return "type3"

    if has_multiline_header(header_row) and any(
        "return" in h.lower() and "b2b" in h.lower() for h in headers.values()
    ):
        if "NO" in codes or pending_se:
            return "type5"

    if (pending_se or codes <= {"SE", "DK"}) and has_multiline_header(header_row):
        if has_paired_weight_rows(weight_rows):
            return "type1"

    if len(codes) == 1:
        return "type2"

    return "type2"


def process_postnord_block(
    country_codes: list[str],
    header_row: dict,
    weight_rows: list[dict[str, str]],
    origin: OriginInfo,
    section_title: str,
    valid_from: str,
    valid_to: str,
    pending_se: bool = False,
) -> list[LaneRow]:
    block_type = classify_postnord_block(
        country_codes, header_row, weight_rows, pending_se
    )

    if block_type == "type1":
        return process_postnord_type1_se_dk(
            header_row, weight_rows, origin, section_title, valid_from, valid_to
        )
    if block_type == "type3":
        return process_postnord_type3_norway_zones(
            header_row, weight_rows, origin, section_title, valid_from, valid_to
        )
    if block_type == "type4":
        return process_postnord_type4_multi_return(
            header_row, weight_rows, origin, section_title, valid_from, valid_to
        )
    if block_type == "type5":
        return process_postnord_type5_return_split(
            header_row, weight_rows, origin, section_title, valid_from, valid_to
        )

    country = country_codes[0] if country_codes else ""
    return process_postnord_type2_single(
        country, header_row, weight_rows, origin, section_title, valid_from, valid_to
    )


def parse_destination_from_header(header: str, fallback: str | None = None) -> str:
    text = header.split("\n")[0]
    low = text.lower()
    for country, code in COUNTRY_TO_CODE.items():
        if country in low:
            return code
    m = re.search(r"\bzone\s+\d+\s+.*?\b([A-Z]{2})\b", text, re.I)
    if m:
        return m.group(1).upper()
    if fallback:
        return fallback
    return ""


def resolve_carrier_partner(section_title: str, partner: str) -> str:
    key = partner.strip().upper()
    if key in PARTNER_ALIASES:
        return PARTNER_ALIASES[key]
    if "poste italian" in section_title.lower():
        return "SDA IT"
    return partner


def _header_line(header: str) -> str:
    return header.split("\n")[0].strip()


def _partner_key(partner: str, section_title: str = "") -> str:
    return resolve_carrier_partner(section_title, partner).upper()


def parse_service_type(
    header: str,
    *,
    partner: str,
    cost_key: str | None = None,
    section_title: str = "",
) -> str:
    text = _header_line(header)
    low = text.lower()
    pk = _partner_key(partner, section_title)
    ck = (cost_key or "").upper()

    if pk in ("POSTE ITALIANE", "SDA IT") or "poste italian" in section_title.lower():
        if ck == "COST1" and "return" in low and "pudo" in low:
            return "RETURN"
        if ck == "COST2" and "return" in low and ("b2b" in low or "home" in low):
            return "RETURN B2B"

    if pk == "POST NL":
        if "return" in low and "netherlands" in low and "pudo" in low:
            return "RETURN"
        if "return" in low and "netherlands" in low and "home" in low:
            return "UNDELIVERED"
        if "return" in low and "belgium" in low and "pudo" in low:
            return "RETURN"

    if pk == "BRT":
        if ck == "COST1" and "home" in low and "pudo" in low:
            return "HOME AND PUDO"
        if ck == "COST2" and "undeliverable" in low:
            return "UNDELIVERED"

    if "MONDIAL RELAY" in pk:
        if ck == "COST1" and "return" in low and "pudo" in low:
            return "RETURN"

    if pk == "DPD FR":
        if ck == "COST2" and "b2c" in low:
            return "B2C"
        if ck == "COST4" and "return" in low and "pudo" in low:
            return "UNDELIVERED PUDO"
        if re.search(r"HOME\s*&\s*Undeliverable", text, re.I):
            return "UNDELIVERED HOME B2B"
        if ck == "COST6" and "return" in low and "b2b" in low:
            return "RETURN"

    if re.search(r"HOME\s*&\s*Undeliverable", text, re.I):
        return "UNDELIVERED HOME B2B"

    for pattern, label in SERVICE_PATTERNS:
        if pattern.search(text):
            if label == "Return":
                for sub_pattern, sub in SERVICE_PATTERNS[1:]:
                    if sub_pattern.search(text):
                        return f"Return {sub}"
                return "Return"
            return label
    return text.strip()


def is_garons_origin(origin: OriginInfo) -> bool:
    return origin.city.lower() == "garons"


def _service_upper(service: str) -> str:
    return service.upper()


def is_return_pudo_service(service: str) -> bool:
    s = _service_upper(service)
    return "RETURN" in s and "PUDO" in s and "UNDELIVERED" not in s


def is_return_home_service(service: str) -> bool:
    s = _service_upper(service)
    if "GARONS" in s:
        return False
    return "RETURN" in s and "HOME" in s and "PUDO" not in s


def is_undelivered_home_service(service: str) -> bool:
    s = _service_upper(service)
    return "UNDELIVERED" in s and "HOME" in s


def lane_cost_fingerprint(lane: LaneRow) -> tuple:
    return (
        lane.carrier,
        lane.destination_country,
        lane.transport_title,
        frozenset(lane.brackets.items()),
    )


def clone_lane(lane: LaneRow, service_type: str) -> LaneRow:
    return LaneRow(
        carrier=lane.carrier,
        carrier_partner=lane.carrier_partner,
        origin_country=lane.origin_country,
        origin_postal=lane.origin_postal,
        destination_country=lane.destination_country,
        service_type=service_type,
        valid_from=lane.valid_from,
        valid_to=lane.valid_to,
        transport_title=lane.transport_title,
        currency=lane.currency,
        brackets=dict(lane.brackets),
        bracket_rate_types=dict(lane.bracket_rate_types),
        forward_filled=set(lane.forward_filled),
    )


def expand_service_lanes(
    lanes: list[LaneRow], origin: OriginInfo, partner: str, section_title: str
) -> list[LaneRow]:
    """Add duplicate rows for carrier-specific undelivered / Garons services."""
    if is_postnord_partner(partner, section_title):
        return lanes

    expanded: list[LaneRow] = []
    garons_duped: set[tuple] = set()
    pk = _partner_key(partner, section_title)
    garons = is_garons_origin(origin)

    for lane in lanes:
        expanded.append(lane)
        svc = lane.service_type

        if pk == "SEUR":
            if is_return_pudo_service(svc):
                expanded.append(clone_lane(lane, "UNDELIVERED PUDO"))
            if is_return_home_service(svc):
                expanded.append(clone_lane(lane, "UNDELIVERED HOME"))

        if garons and (
            is_return_home_service(svc) or is_undelivered_home_service(svc)
        ):
            fp = lane_cost_fingerprint(lane)
            if fp not in garons_duped:
                garons_duped.add(fp)
                expanded.append(clone_lane(lane, "RETURN HOME GARONS"))

    return expanded


def parse_currency(header: str, cost_sample: str | None = None) -> str:
    for text in (header, cost_sample or ""):
        for sym, code in CURRENCY_SYMBOLS.items():
            if sym in text:
                return code
    return "EUR"


def parse_price_validity(text: str) -> tuple[str, str]:
    m = re.search(
        r"from\s+(\d{1,2}/\d{1,2}/\d{4})\s+until\s+(\d{1,2}/\d{1,2}/\d{4})",
        text,
        re.I,
    )
    if not m:
        return "", ""

    def to_display(date_str: str) -> str:
        d, mo, y = date_str.split("/")
        return f"{int(d):02d}.{int(mo):02d}.{y}"

    return to_display(m.group(1)), to_display(m.group(2))


def parse_weight_label(weight: str) -> tuple[str, float | None]:
    w = weight.strip()
    if re.search(r"add\.?\s*kg", w, re.I):
        return "adder", None
    m = re.match(r"([\d.,]+)\s*kg", w, re.I)
    if m:
        val = float(m.group(1).replace(",", "."))
        if val == int(val):
            return f"<= {int(val)}", val
        return f"<= {val}", val
    return w, None


def parse_cost_value(raw: str | None) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s == "-":
        return None
    s = s.replace("€", "").replace(",", ".").strip()
    s = s.lstrip("+")
    try:
        return float(s)
    except ValueError:
        return None


def forward_fill_brackets(
    weight_rows: list[dict[str, str]], cost_key: str
) -> list[tuple[str, float | None, str, bool]]:
    parsed: list[tuple[str, float | None, float | None, str]] = []
    for row in weight_rows:
        label, numeric = parse_weight_label(row["Weight"])
        parsed.append(
            (label, numeric, parse_cost_value(row_get_cost(row, cost_key)), row["Weight"])
        )

    filled: list[tuple[str, float | None, str, bool]] = []
    for i, (label, _numeric, cost, raw_weight) in enumerate(parsed):
        forward_filled = False
        if cost is None:
            for j in range(i + 1, len(parsed)):
                if parsed[j][2] is not None:
                    cost = parsed[j][2]
                    forward_filled = True
                    break
        filled.append((label, cost, raw_weight, forward_filled))
    return filled


def block_max_weight_kg(weight_rows: list[dict[str, str]]) -> int | None:
    max_kg: float | None = None
    for row in weight_rows:
        label, numeric = parse_weight_label(row["Weight"])
        if label != "adder" and numeric is not None:
            max_kg = numeric if max_kg is None else max(max_kg, numeric)
    return int(max_kg) if max_kg is not None else None


def lane_weight_profile(
    weight_rows: list[dict[str, str]], cost_key: str
) -> tuple[list[str], dict[str, str], int | None, bool]:
    """Brackets and transport title max weight for one service column (lane)."""
    filled = forward_fill_brackets(weight_rows, cost_key)
    block_max = block_max_weight_kg(weight_rows)

    labels: list[str] = []
    rate_types: dict[str, str] = {}
    max_kg_with_cost: float | None = None
    has_adder = False

    for label, cost, _raw_weight, _forward_filled in filled:
        if label == "adder":
            if cost is not None:
                has_adder = True
            continue
        if cost is None:
            continue
        if label not in labels:
            labels.append(label)
            rate_types[label] = "Flat"
        _, numeric = parse_weight_label(_raw_weight)
        if numeric is not None:
            max_kg_with_cost = (
                numeric if max_kg_with_cost is None else max(max_kg_with_cost, numeric)
            )

    max_kg_int = int(max_kg_with_cost) if max_kg_with_cost is not None else None

    if has_adder:
        threshold = block_max if block_max is not None else (max_kg_int or 30)
        adder_label = f">{threshold}"
        labels.append(adder_label)
        rate_types[adder_label] = "p/unit"
        max_kg_int = threshold

    return labels, rate_types, max_kg_int, has_adder


def lane_bracket_values(
    weight_rows: list[dict[str, str]],
    cost_key: str,
    bracket_labels: list[str],
) -> tuple[dict[str, object], set[str]]:
    filled = forward_fill_brackets(weight_rows, cost_key)
    by_label = {label: (cost, is_ff) for label, cost, _raw, is_ff in filled}

    values: dict[str, object] = {}
    forward_filled_cols: set[str] = set()
    for col in bracket_labels:
        if col.startswith(">"):
            entry = by_label.get("adder")
        else:
            entry = by_label.get(col)
        if entry is None or entry[0] is None:
            continue
        values[col] = entry[0]
        if entry[1]:
            forward_filled_cols.add(col)
    return values, forward_filled_cols


def build_transport_title(
    partner: str, origin: OriginInfo, max_kg: int | None, has_adder: bool
) -> str:
    kg = max_kg if max_kg is not None else 30
    if has_adder:
        return f"Transport Cost ({partner} {origin.label} {kg}kg + add kg)"
    return f"Transport Cost ({partner} {origin.label} {kg} kg)"


def build_carrier(partner: str, origin: OriginInfo, destination: str) -> str:
    return f"{partner} {origin.city} {destination}".strip()


def extract_headers(row: dict) -> dict[str, str]:
    headers: dict[str, str] = {}
    for key in row:
        cost_key = normalize_cost_key(key)
        if not cost_key.startswith("Cost"):
            continue
        val = row.get(key)
        if val is None:
            continue
        text = str(val).strip()
        if not text or text.lower() == "weight (up to)":
            continue
        if "€" in text or "shipping" in text.lower() or "return" in text.lower():
            headers[cost_key] = text
    return headers


def block_has_prices(block: PriceBlock) -> bool:
    if not block.headers or not block.weight_rows:
        return False
    for key in block.headers:
        for row in block.weight_rows:
            if parse_cost_value(row_get_cost(row, key)) is not None:
                return True
    return False


def lanes_from_block(
    block: PriceBlock, valid_from: str, valid_to: str
) -> list[LaneRow]:
    lanes: list[LaneRow] = []

    for cost_key, header in block.headers.items():
        header_line = header.split("\n")[0].strip()
        if "discount" in header_line.lower():
            continue

        dest = (
            block.destination_code
            or parse_destination_from_header(header, block.section_country)
        )
        if not dest:
            continue

        bracket_labels, rate_types, max_kg, has_adder = lane_weight_profile(
            block.weight_rows, cost_key
        )
        if not bracket_labels:
            continue

        brackets, forward_filled = lane_bracket_values(
            block.weight_rows, cost_key, bracket_labels
        )
        if not brackets:
            continue

        display_partner = resolve_carrier_partner(
            block.section_title, block.partner
        )
        service = parse_service_type(
            header,
            partner=block.partner,
            cost_key=cost_key,
            section_title=block.section_title,
        )
        transport_title = build_transport_title(
            block.partner, block.origin, max_kg, has_adder
        )
        sample_cost = next(iter(brackets.values()))
        currency = parse_currency(header, str(sample_cost) if sample_cost else None)

        lane = LaneRow(
            carrier=build_carrier(display_partner, block.origin, dest),
            carrier_partner=display_partner,
            origin_country=block.origin.country,
            origin_postal=block.origin.postal_code,
            destination_country=dest,
            service_type=service,
            valid_from=valid_from,
            valid_to=valid_to,
            transport_title=transport_title,
            currency=currency,
            brackets=brackets,
            bracket_rate_types=rate_types,
            forward_filled=forward_filled,
        )
        lanes.append(lane)
    return expand_service_lanes(
        lanes, block.origin, block.partner, block.section_title
    )


def parse_main_costs(
    main_costs: list[dict],
    price_validity: str,
    carrier_label: str = "Seven Senders",
) -> list[LaneRow]:
    valid_from, valid_to = parse_price_validity(price_validity)
    lanes: list[LaneRow] = []

    partner = ""
    section_title = ""
    section_country: str | None = None
    origin: OriginInfo | None = None
    country_codes: list[str] = []
    header: dict[str, str] | None = None
    weight_rows: list[dict[str, str]] = []
    postnord_pending_se = False

    in_packing_density = False
    skip_floater_garons = False
    floater_origin: OriginInfo | None = None
    floater_pending_sp = False
    floater_merged_weights: list[dict[str, str]] = []
    floater_region_weights: list[dict[str, str]] = []
    floater_region_name: str | None = None

    def flush_floater_region() -> None:
        nonlocal floater_region_weights, floater_region_name, floater_merged_weights
        if floater_region_name and floater_region_weights and floater_origin:
            lanes.extend(
                process_floater_single_region(
                    floater_region_name,
                    floater_region_weights,
                    floater_origin,
                    carrier_label,
                    valid_from,
                    valid_to,
                )
            )
        floater_region_weights = []
        floater_region_name = None

    def flush_floater_merged() -> None:
        nonlocal floater_merged_weights, floater_pending_sp
        if floater_merged_weights and floater_origin and floater_pending_sp:
            lanes.extend(
                process_floater_merged_sp_it(
                    floater_merged_weights,
                    floater_origin,
                    carrier_label,
                    valid_from,
                    valid_to,
                )
            )
        floater_merged_weights = []
        floater_pending_sp = False

    def flush_block() -> None:
        nonlocal header, weight_rows, origin, country_codes, postnord_pending_se
        if not partner or not origin or not header or not weight_rows:
            header = None
            weight_rows = []
            postnord_pending_se = False
            return

        if is_postnord_partner(partner, section_title):
            block_countries = list(country_codes)
            if postnord_pending_se and "DK" in block_countries and "SE" not in block_countries:
                block_countries = ["SE", "DK"]
            lanes.extend(
                process_postnord_block(
                    block_countries,
                    header,
                    weight_rows,
                    origin,
                    section_title,
                    valid_from,
                    valid_to,
                    pending_se=postnord_pending_se,
                )
            )
        else:
            dest = country_codes[0] if len(country_codes) == 1 else None
            block = PriceBlock(
                partner=partner,
                section_title=section_title,
                section_country=section_country,
                origin=origin,
                destination_code=dest,
                headers=header,
                weight_rows=weight_rows,
            )
            if block_has_prices(block):
                if len(country_codes) > 1 and not is_postnord_partner(partner, section_title):
                    for code in country_codes:
                        block_copy = PriceBlock(
                            partner=partner,
                            section_title=section_title,
                            section_country=section_country,
                            origin=origin,
                            destination_code=code,
                            headers=header,
                            weight_rows=weight_rows,
                        )
                        lanes.extend(lanes_from_block(block_copy, valid_from, valid_to))
                else:
                    lanes.extend(lanes_from_block(block, valid_from, valid_to))

        header = None
        weight_rows = []
        postnord_pending_se = False

    for row in main_costs:
        if "RateName" in row:
            name = row["RateName"]

            if is_packing_density_floater_start(name):
                flush_block()
                flush_floater_region()
                flush_floater_merged()
                in_packing_density = True
                skip_floater_garons = False
                floater_origin = None
                continue

            if in_packing_density and is_origin_rate_name(name):
                if "garons" in name.lower():
                    flush_floater_region()
                    flush_floater_merged()
                    skip_floater_garons = True
                    continue
                parsed = parse_origin(name, row)
                if parsed:
                    floater_origin = parsed
                    skip_floater_garons = False
                continue

            if in_packing_density and skip_floater_garons:
                continue

            if in_packing_density and is_floater_region_name(name):
                region = name.strip()
                if (
                    region.lower() == "spain & portugal"
                    and not is_floater_header_row(row)
                ):
                    flush_floater_region()
                    flush_floater_merged()
                    floater_pending_sp = True
                    continue
                if (
                    is_floater_header_row(row)
                    and floater_pending_sp
                    and region.lower() == "italy"
                ):
                    floater_merged_weights = []
                    continue
                flush_floater_region()
                flush_floater_merged()
                if is_floater_header_row(row):
                    floater_region_name = region
                    floater_region_weights = []
                continue

            if is_basic_fees(name):
                flush_block()
                flush_floater_region()
                flush_floater_merged()
                in_packing_density = False
                skip_floater_garons = False
                section_title = name.strip()
                partner = parse_partner_from_section(section_title)
                section_country = parse_section_country(section_title)
                country_codes = []
                if is_price_header_row(row):
                    parsed_origin = parse_origin(name, row)
                    if parsed_origin:
                        origin = parsed_origin
                    header = extract_headers(row)
                    weight_rows = []
                elif is_origin_rate_name(name):
                    parsed_origin = parse_origin(name, row)
                    if parsed_origin:
                        origin = parsed_origin
                continue

            if should_skip_section(name):
                flush_block()
                partner = ""
                section_title = ""
                section_country = None
                origin = None
                country_codes = []
                header = None
                weight_rows = []
                continue

            if is_country_rate_name(name):
                codes = parse_country_codes_from_rate_name(name)
                if is_postnord_partner(partner, section_title):
                    if not is_price_header_row(row):
                        if codes == ["SE"]:
                            postnord_pending_se = True
                            continue
                        if set(codes) == {"SE", "DK"} or (
                            len(codes) == 2 and "SE" in codes and "DK" in codes
                        ):
                            postnord_pending_se = True
                            continue
                    flush_block()
                    country_codes = codes
                    if postnord_pending_se and "DK" in country_codes and "SE" not in country_codes:
                        country_codes = ["SE", "DK"]
                    if is_price_header_row(row) and origin:
                        header = row
                        weight_rows = []
                    continue
                flush_block()
                country_codes = codes
                if is_price_header_row(row) and origin:
                    header = extract_headers(row)
                    weight_rows = []
                continue

            if is_origin_rate_name(name):
                flush_block()
                parsed_origin = parse_origin(name, row)
                if parsed_origin:
                    origin = parsed_origin
                if is_price_header_row(row):
                    header = extract_headers(row)
                    weight_rows = []
                continue

            continue

        if is_price_header_row(row):
            flush_block()
            header = extract_headers(row)
            weight_rows = []
            continue

        weight = row.get("Weight")
        if in_packing_density and not skip_floater_garons and is_floater_weight_row(row):
            if floater_pending_sp:
                floater_merged_weights.append(row)
            elif floater_region_name:
                floater_region_weights.append(row)
            continue

        if weight and not is_noise_weight(weight):
            if header is not None:
                weight_rows.append(row)

    flush_block()
    flush_floater_region()
    flush_floater_merged()
    return lanes


def bracket_sort_key(label: str) -> tuple:
    m = re.match(r"<=\s*([\d.]+)", label)
    if m:
        return (0, float(m.group(1)))
    m = re.match(r"<\s*([\d.]+)", label)
    if m:
        return (0, float(m.group(1)))
    m = re.match(r"[≥>=]\s*([\d.]+)", label)
    if m:
        return (1, float(m.group(1)))
    m = re.match(r">\s*([\d.]+)", label)
    if m:
        return (2, float(m.group(1)))
    return (3, label)


def block_column_key(lane: LaneRow) -> str:
    return lane.transport_title


def sample_lane_rate_by(lane: LaneRow) -> str:
    return lane.rate_by_label or DEFAULT_RATE_BY_WEIGHT


def group_lanes_for_columns(lanes: list[LaneRow]) -> list[tuple[str, list[str], list[LaneRow]]]:
    groups: dict[tuple, list[LaneRow]] = {}
    order: list[tuple] = []
    for lane in lanes:
        key = block_column_key(lane)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(lane)

    result = []
    for key in order:
        sample = groups[key][0]
        weights = sorted(sample.brackets.keys(), key=bracket_sort_key)
        result.append((sample.transport_title, weights, groups[key]))
    return result


def _style_header_cell(cell, header_fill, header_font, header_alignment) -> None:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment


def _write_transport_block_headers(
    ws,
    start_col: int,
    title: str,
    weights: list[str],
    sample_lane: LaneRow,
    header_fill,
    header_font,
    header_alignment,
) -> int:
    col = start_col
    rate_by = sample_lane_rate_by(sample_lane)
    cell = ws.cell(row=2, column=col, value=f"Rate by: {rate_by}")
    _style_header_cell(cell, header_fill, header_font, header_alignment)
    col += 1
    for _ in weights:
        cell = ws.cell(row=2, column=col, value="")
        _style_header_cell(cell, header_fill, header_font, header_alignment)
        col += 1

    ws.cell(row=3, column=start_col, value="")
    ws.cell(row=3, column=start_col).fill = header_fill
    col = start_col + 1
    for weight_label in weights:
        cell = ws.cell(row=3, column=col, value=weight_label)
        _style_header_cell(cell, header_fill, header_font, header_alignment)
        col += 1

    cell = ws.cell(row=4, column=start_col, value="Currency")
    _style_header_cell(cell, header_fill, header_font, header_alignment)
    for w_idx, weight_label in enumerate(weights):
        rate_type = sample_lane.bracket_rate_types.get(
            weight_label, "p/unit" if weight_label.startswith(">") else "Flat"
        )
        cell = ws.cell(row=4, column=start_col + 1 + w_idx, value=rate_type)
        _style_header_cell(cell, header_fill, header_font, header_alignment)

    end_col = start_col + len(weights)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=title)
    _style_header_cell(cell, header_fill, header_font, header_alignment)
    return end_col


def _write_floater_block_headers(
    ws,
    start_col: int,
    floater: FloaterColumn,
    weights: list[str],
    header_fill,
    header_font,
    header_alignment,
) -> int:
    end_col = start_col + len(weights)

    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=floater.title)
    _style_header_cell(cell, header_fill, header_font, header_alignment)

    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    cell = ws.cell(row=2, column=start_col, value=floater.applies_if)
    _style_header_cell(cell, header_fill, header_font, header_alignment)

    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
    cell = ws.cell(row=3, column=start_col, value=f"Rate by: {floater.rate_by}")
    _style_header_cell(cell, header_fill, header_font, header_alignment)

    ws.cell(row=4, column=start_col, value="")
    ws.cell(row=4, column=start_col).fill = header_fill
    col = start_col + 1
    for weight_label in weights:
        cell = ws.cell(row=4, column=col, value=weight_label)
        _style_header_cell(cell, header_fill, header_font, header_alignment)
        col += 1

    cell = ws.cell(row=5, column=start_col, value="Currency")
    _style_header_cell(cell, header_fill, header_font, header_alignment)
    for w_idx, weight_label in enumerate(weights):
        rate_type = floater.bracket_rate_types.get(
            weight_label, "p/unit" if weight_label.startswith(">") else "Flat"
        )
        cell = ws.cell(row=5, column=start_col + 1 + w_idx, value=rate_type)
        _style_header_cell(cell, header_fill, header_font, header_alignment)

    return end_col


def write_rate_card_sheet(
    workbook: Workbook,
    lanes: list[LaneRow],
    sheet_title: str = SHEET_TITLE,
    floater_columns: list[FloaterColumn] | None = None,
) -> None:
    ws = workbook.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    forward_fill_highlight = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )

    transport_lanes, parsed_floaters = split_transport_and_floater(lanes)
    floaters = floater_columns if floater_columns is not None else parsed_floaters

    num_fixed = len(FIXED_COLS)
    column_groups = group_lanes_for_columns(transport_lanes)
    header_row_count = 6 if floaters else 5
    data_start_row = header_row_count

    for col_idx, name in enumerate(FIXED_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        _style_header_cell(cell, header_fill, header_font, header_alignment)

    col = num_fixed + 1
    transport_blocks: list[tuple[int, list[str], str]] = []

    for title, weights, group_lanes in column_groups:
        start_col = col
        end_col = _write_transport_block_headers(
            ws,
            start_col,
            title,
            weights,
            group_lanes[0],
            header_fill,
            header_font,
            header_alignment,
        )
        transport_blocks.append((start_col, weights, title))
        col = end_col + 1

    floater_blocks: list[tuple[int, list[str], FloaterColumn]] = []
    for floater in floaters:
        weights = sorted(floater.brackets.keys(), key=bracket_sort_key)
        start_col = col
        end_col = _write_floater_block_headers(
            ws, start_col, floater, weights, header_fill, header_font, header_alignment
        )
        floater_blocks.append((start_col, weights, floater))
        col = end_col + 1

    for row_idx in range(2, header_row_count):
        for c in range(1, num_fixed + 1):
            cell = ws.cell(row=row_idx, column=c, value="")
            cell.fill = header_fill

    transport_block_map: dict[tuple, tuple[int, list[str]]] = {
        (title, tuple(weights)): (start_col, weights)
        for start_col, weights, title in transport_blocks
    }

    transport_rows: list[tuple[LaneRow, tuple]] = []
    for title, weights, group_lanes in column_groups:
        block_key = (title, tuple(weights))
        for lane in group_lanes:
            transport_rows.append((lane, block_key))

    data_row = data_start_row
    for lane, block_key in transport_rows:
        col = 1
        for value in (
            lane.carrier,
            lane.carrier_partner,
            lane.origin_country,
            lane.origin_postal,
            lane.destination_country,
            lane.service_type,
            lane.valid_from,
            lane.valid_to,
        ):
            ws.cell(row=data_row, column=col, value=value)
            col += 1

        start_col, bracket_labels = transport_block_map[block_key]
        ws.cell(row=data_row, column=start_col, value=lane.currency)
        for w_idx, weight_label in enumerate(bracket_labels):
            cell = ws.cell(
                row=data_row,
                column=start_col + 1 + w_idx,
                value=lane.brackets.get(weight_label, ""),
            )
            cell.alignment = Alignment(horizontal="center")
            if weight_label in lane.forward_filled:
                cell.fill = forward_fill_highlight

        for floater_start, floater_weights, floater in floater_blocks:
            ws.cell(row=data_row, column=floater_start, value=floater.currency)
            for w_idx, weight_label in enumerate(floater_weights):
                cell = ws.cell(
                    row=data_row,
                    column=floater_start + 1 + w_idx,
                    value=floater.brackets.get(weight_label, ""),
                )
                cell.alignment = Alignment(horizontal="center")
                if weight_label in floater.forward_filled:
                    cell.fill = forward_fill_highlight

        data_row += 1

    total_cols = col - 1 if transport_rows or floaters else num_fixed
    last_row = data_row - 1
    for c in range(1, total_cols + 1):
        letter = get_column_letter(c)
        max_len = 10
        for r in range(1, min(last_row + 1, 54)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    ws.freeze_panes = f"A{data_start_row}"
    if transport_rows and last_row >= data_start_row:
        filter_row = data_start_row - 1
        ws.auto_filter.ref = (
            f"A{filter_row}:{get_column_letter(total_cols)}{last_row}"
        )


def transform_json_to_xlsx(data: dict, out_path: Path) -> Path:
    fields = data.get("fields") or {}
    main_costs = fields.get("MainCosts") or data.get("MainCosts") or []
    price_validity = fields.get("PriceValidity") or data.get("PriceValidity") or ""

    carrier_label = normalize_carrier_label(
        fields.get("Carrier") or data.get("Carrier") or "Seven Senders"
    )
    lanes = parse_main_costs(main_costs, price_validity, carrier_label)
    transport_lanes, floater_columns = split_transport_and_floater(lanes)
    if not transport_lanes and not floater_columns:
        raise ValueError("No MainCosts lanes could be built from the input JSON.")

    wb = Workbook()
    write_rate_card_sheet(wb, transport_lanes, floater_columns=floater_columns)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def process_file(input_path: Path, output_path: Path | None = None) -> Path:
    with input_path.open(encoding="utf-8") as f:
        data = json.load(f)

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / (input_path.stem + ".xlsx")

    return transform_json_to_xlsx(data, output_path)


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    if args:
        input_path = Path(args[0])
        output_path = Path(args[1]) if len(args) > 1 else None
    else:
        json_files = sorted(PROCESSING_DIR.glob("*.json"))
        if not json_files:
            print(f"No JSON files in {PROCESSING_DIR}", file=sys.stderr)
            return 1
        input_path = json_files[0]
        output_path = None

    out = process_file(input_path, output_path)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
